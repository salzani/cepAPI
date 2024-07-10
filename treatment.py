import requests
import re
import json
from openpyxl import *
from datetime import datetime




apiURL = 'https://viacep.com.br/ws'

def dataFormat(jsonData):
    cityInfo = jsonData.get('localidade')
    ufInfo = jsonData.get('uf')
    ibgeInfo = jsonData.get('ibge')
    dddInfo = jsonData.get('ddd')

    print('')
    print('||||||||||||||||||||||||')
    print(f'cidade: {cityInfo}')
    print(f'UF: {ufInfo}')
    print(f'IBGE: {ibgeInfo}')
    print(f'dddInfo: 0{dddInfo}')
    print('||||||||||||||||||||||||')

    return {
        'Cidade': cityInfo,
        'UF': ufInfo,
        'IBGE': ibgeInfo,
        'DDD': f'0{dddInfo}'
    }


def returnCep(apiReceive):
    apiResponse = requests.get(apiReceive)

    if apiResponse.status_code == 200:
        data = apiResponse.json()
        return dataFormat(data)
    else:
        print(f'Error to request | STATUS: {apiResponse.status_code}')
        return None


def findNextEmptyRow(sheet):
    for row in range(2, sheet.max_row + 2):
        if sheet[f'A{row}'].value is None:
            return row
    return sheet.max_row + 1


### DATA E HORA
dateNow = datetime.now()

formattedDateNow = dateNow.strftime("%Y-%m-%d %H:%M:%S")
### DATA E HORA


cepInput = input('Entre com o CEP: ')

cep = re.sub(r'[^a-zA-Z0-9]', '', cepInput)

cepTreatment = f'{apiURL}/{cep}/json/'

finalData = returnCep(cepTreatment)


print('')
print('Preencher planilha?')
choice = input('(S) | (N)\n').lower()
print('')


if choice == 's':
    # Carrega a planilha existente
    workbook = load_workbook('fill.xlsx')
    sheet = workbook.active

    # Encontra a próxima linha vazia
    next_row = findNextEmptyRow(sheet)

    # Preenche as colunas "Data e hora", "Cidade", "IBGE", "UF" e "DDD" com os dados obtidos.
    sheet[f'A{next_row}'] = formattedDateNow
    sheet[f'B{next_row}'] = finalData['Cidade']
    sheet[f'C{next_row}'] = finalData['IBGE']
    sheet[f'D{next_row}'] = finalData['UF']
    sheet[f'E{next_row}'] = finalData['DDD']

    # Salva a planilha
    workbook.save('fill.xlsx')
    print('Planilha preenchida.')
else:
    print('Finalizado.')
